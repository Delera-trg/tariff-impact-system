# -*- coding: utf-8 -*-
"""
Excel导出模块 - 导出关税冲击分析结果
使用openpyxl库生成Excel报告
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, List, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("请安装openpyxl: pip install openpyxl")
    raise


class Exporter:
    """
    Excel导出器

    将关税冲击分析结果导出为Excel文件
    """

    def __init__(self, output_dir: str = "exports"):
        """
        初始化导出器

        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export_to_excel(
        self,
        calculation_result: Dict[str, Any],
        sensitivity_result: Dict[str, Any] = None,
        comparison_result: Dict[str, Any] = None,
        file_name: str = None
    ) -> str:
        """
        导出计算结果到Excel

        Args:
            calculation_result: 基本计算结果
            sensitivity_result: 敏感性分析结果
            comparison_result: 行业对比结果
            file_name: 文件名（不含扩展名）

        Returns:
            str: 导出文件的完整路径
        """
        # 生成文件名
        if file_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"tariff_analysis_{timestamp}"

        file_path = os.path.join(self.output_dir, f"{file_name}.xlsx")

        # 创建工作簿
        wb = Workbook()

        # 移除默认sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 1. 基本信息表
        self._create_summary_sheet(wb, calculation_result)

        # 2. 价格变化表
        self._create_price_sheet(wb, calculation_result)

        # 3. 福利效应表
        self._create_welfare_sheet(wb, calculation_result)

        # 4. 敏感性分析表（如果有）
        if sensitivity_result:
            self._create_sensitivity_sheet(wb, sensitivity_result)

        # 5. 行业对比表（如果有）
        if comparison_result:
            self._create_comparison_sheet(wb, comparison_result)

        # 保存文件
        wb.save(file_path)
        return file_path

    def _create_summary_sheet(self, wb: Workbook, result: Dict[str, Any]) -> None:
        """创建基本信息表"""
        ws = wb.create_sheet("基本信息")

        # 样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 标题
        ws['A1'] = "中国商品价格传导与关税冲击分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:D2')

        # 行业信息
        ws['A4'] = "行业信息"
        ws['A4'].font = header_font

        industry = result.get("industry", {})
        data = [
            ("HS编码", industry.get("hs_code", "")),
            ("行业名称", industry.get("name", "")),
            ("分类", industry.get("category", "")),
            ("计量单位", industry.get("unit", "")),
        ]

        for i, (key, value) in enumerate(data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value

        # 计算参数
        ws['A10'] = "计算参数"
        ws['A10'].font = header_font

        params = result.get("params", {})
        # 安全获取参数，处理None情况
        tariff_rate = params.get('tariff_rate') or 0
        base_price = params.get('base_price') or 0
        pass_through_1 = params.get('pass_through_1') or 0
        pass_through_2 = params.get('pass_through_2') or 0
        elasticity = params.get('elasticity') or 0

        data = [
            ("关税税率", f"{tariff_rate*100:.1f}%"),
            ("基准进口价格", f"¥{base_price:,.2f}"),
            ("进口→批发传导率", f"{pass_through_1*100:.0f}%"),
            ("批发→零售传导率", f"{pass_through_2*100:.0f}%"),
            ("需求弹性", f"{elasticity:.2f}"),
        ]

        for i, (key, value) in enumerate(data, start=11):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

    def _create_price_sheet(self, wb: Workbook, result: Dict[str, Any]) -> None:
        """创建价格变化表"""
        ws = wb.create_sheet("价格变化")

        # 样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")

        # 表头
        headers = ["环节", "税前价格", "税后价格", "变化额", "变化率"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 数据
        price_changes = result.get("price_changes", {})
        stages = ["import", "wholesale", "retail"]
        stage_names = {"import": "进口", "wholesale": "批发", "retail": "零售"}

        for row, stage in enumerate(stages, start=2):
            stage_data = price_changes.get(stage, {})
            ws.cell(row=row, column=1, value=stage_names.get(stage, stage))
            ws.cell(row=row, column=2, value=stage_data.get("before", 0))
            ws.cell(row=row, column=3, value=stage_data.get("after", 0))
            ws.cell(row=row, column=4, value=stage_data.get("change", 0))
            ws.cell(row=row, column=5, value=f"{stage_data.get('change_rate', 0):.2f}%")

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15

    def _create_welfare_sheet(self, wb: Workbook, result: Dict[str, Any]) -> None:
        """创建福利效应表"""
        ws = wb.create_sheet("福利效应")

        # 样式
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")

        # 表头
        headers = ["项目", "金额 (元)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # 数据
        welfare = result.get("welfare_effects", {})
        data = [
            ("消费者剩余变化", welfare.get("consumer_surplus_change", 0)),
            ("生产者剩余变化", welfare.get("producer_surplus_change", 0)),
            ("政府关税收入", welfare.get("government_revenue", 0)),
            ("无谓损失 (DWL)", welfare.get("deadweight_loss", 0)),
        ]

        for row, (key, value) in enumerate(data, start=2):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_sensitivity_sheet(self, wb: Workbook, result: Dict[str, Any]) -> None:
        """创建敏感性分析表"""
        ws = wb.create_sheet("敏感性分析")

        # 样式
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")

        # 表头
        headers = ["关税税率", "进口价格", "批发价格", "零售价格", "零售价变化率", "政府收入", "无谓损失"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # 数据
        analysis = result.get("analysis", [])
        for row, item in enumerate(analysis, start=2):
            ws.cell(row=row, column=1, value=f"{item.get('tariff_rate_percent', 0):.0f}%")
            ws.cell(row=row, column=2, value=item.get("import_price", 0))
            ws.cell(row=row, column=3, value=item.get("wholesale_price", 0))
            ws.cell(row=row, column=4, value=item.get("retail_price", 0))
            ws.cell(row=row, column=5, value=f"{item.get('retail_price_change_rate', 0):.2f}%")
            ws.cell(row=row, column=6, value=item.get("government_revenue", 0))
            ws.cell(row=row, column=7, value=item.get("deadweight_loss", 0))

        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15

    def _create_comparison_sheet(self, wb: Workbook, result: Dict[str, Any]) -> None:
        """创建行业对比表"""
        ws = wb.create_sheet("行业对比")

        # 样式
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")

        # 表头
        headers = ["行业名称", "HS编码", "基准价格", "零售价", "变化率", "政府收入", "无谓损失"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # 数据
        comparison = result.get("comparison", [])
        for row, item in enumerate(comparison, start=2):
            ws.cell(row=row, column=1, value=item.get("name", ""))
            ws.cell(row=row, column=2, value=item.get("hs_code", ""))
            ws.cell(row=row, column=3, value=item.get("base_price", 0))
            ws.cell(row=row, column=4, value=item.get("retail_price_after", 0))
            ws.cell(row=row, column=5, value=f"{item.get('retail_price_change_rate', 0):.2f}%")
            ws.cell(row=row, column=6, value=item.get("government_revenue", 0))
            ws.cell(row=row, column=7, value=item.get("deadweight_loss", 0))

        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15

    def export_to_csv(self, result: Dict[str, Any], file_name: str = None) -> str:
        """
        导出为CSV格式

        Args:
            result: 计算结果
            file_name: 文件名

        Returns:
            str: 导出文件的路径
        """
        if file_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"tariff_analysis_{timestamp}.csv"

        file_path = os.path.join(self.output_dir, file_name)

        # 创建DataFrame
        price_changes = result.get("price_changes", {})
        data = []

        for stage, stage_name in [("import", "进口"), ("wholesale", "批发"), ("retail", "零售")]:
            stage_data = price_changes.get(stage, {})
            data.append({
                "环节": stage_name,
                "税前价格": stage_data.get("before", 0),
                "税后价格": stage_data.get("after", 0),
                "变化额": stage_data.get("change", 0),
                "变化率": f"{stage_data.get('change_rate', 0):.2f}%"
            })

        df = pd.DataFrame(data)
        df.to_csv(file_path, index=False, encoding='utf-8-sig')

        return file_path


# 测试代码
if __name__ == "__main__":
    import sys
    sys.path.insert(0, ".")
    from modules.calculator import TariffCalculator

    # 创建计算器
    calc = TariffCalculator()

    # 基本计算
    print("计算中...")
    result = calc.calculate(hs_code="0101.10", tariff_rate=0.10)
    print(f"计算结果: {result.get('success')}")

    # 敏感性分析
    sensitivity = calc.sensitivity_analysis(hs_code="0101.10")
    print(f"敏感性分析: {sensitivity.get('success')}")

    # 导出
    exporter = Exporter()
    file_path = exporter.export_to_excel(result, sensitivity_result=sensitivity)
    print(f"已导出到: {file_path}")
